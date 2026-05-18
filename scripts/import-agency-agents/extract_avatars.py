#!/usr/bin/env python3
"""Extract embedded avatar images from 智能体目录.xlsx, output as PNG files keyed by agent name."""

import sys
import os
import re
from pathlib import Path
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image

EXCEL_PATH = Path("/Users/zongkelong/workspace/agency-agents-zh/智能体目录.xlsx")
OUTPUT_DIR = Path("/Users/zongkelong/workspace/nexu/.worktree/tabby-ui/scripts/import-agency-agents/extracted_avatars")

# Map directory name to Chinese department name
DEPT_MAP = {
    "academic": "学术部",
    "design": "设计部",
    "engineering": "工程部",
    "finance": "金融部",
    "game-development": "游戏开发部",
    "hr": "人力资源部",
    "legal": "法务部",
    "marketing": "营销部",
    "paid-media": "付费媒体部",
    "product": "产品部",
    "project-management": "项目管理部",
    "sales": "销售部",
    "spatial-computing": "空间计算部",
    "specialized": "专项部",
    "supply-chain": "供应链部",
    "support": "支持部",
    "testing": "测试部",
}


def slugify(name):
    """Derive slug from filename in the Excel's file path column."""
    # The file path looks like "marketing/douyin-strategist.md"
    # Return the filename without .md extension
    return name


def main():
    print("Loading Excel (this may take a minute for 721MB file)...")
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, rows: {ws.max_row}")

    # Read agent names from column C and file paths from column E
    # Row 1-2: headers, Row 3: column headers, Row 4+: data
    agents = []  # list of (name, filepath, dept)
    for row in range(4, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=3).value
        filepath_cell = ws.cell(row=row, column=5).value
        if not name_cell:
            continue
        name = str(name_cell).strip()
        filepath = str(filepath_cell).strip() if filepath_cell else ""
        # Determine department from filepath
        dept = "专项部"
        if filepath:
            top_dir = filepath.split("/")[0]
            dept = DEPT_MAP.get(top_dir, "专项部")
        agents.append((name, filepath, dept, row))

    print(f"Found {len(agents)} agents in spreadsheet")

    # Extract images: openpyxl stores them in ws._images
    images = ws._images
    print(f"Found {len(images)} embedded images")

    # Match images to rows by anchor row
    row_to_image = {}
    for img in images:
        # The anchor tells us which row the image is in
        if hasattr(img, "anchor") and hasattr(img.anchor, "_from"):
            anchor_row = img.anchor._from.row + 1  # openpyxl is 0-indexed
            row_to_image[anchor_row] = img

    print(f"Matched {len(row_to_image)} images to rows")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    extracted = 0
    missing = 0

    for name, filepath, dept, row in agents:
        # Derive slug from filepath
        slug = Path(filepath).stem if filepath else ""

        if row in row_to_image:
            img = row_to_image[row]
            # Get image data
            if hasattr(img, "ref"):
                # For images loaded from file
                try:
                    pil_img = Image.open(img.ref)
                    out_path = OUTPUT_DIR / f"{slug}.png"
                    pil_img.save(out_path, "PNG")
                    size_kb = os.path.getsize(out_path) / 1024
                    print(f"  OK {slug}.png ({size_kb:.0f} KB) — {name} [{dept}]")
                    extracted += 1
                except Exception as e:
                    print(f"  ERR {slug}: {e}")
                    missing += 1
            elif hasattr(img, "_data"):
                # For images embedded as bytes
                try:
                    pil_img = Image.open(BytesIO(img._data()))
                    out_path = OUTPUT_DIR / f"{slug}.png"
                    pil_img.save(out_path, "PNG")
                    size_kb = os.path.getsize(out_path) / 1024
                    print(f"  OK {slug}.png ({size_kb:.0f} KB) — {name} [{dept}]")
                    extracted += 1
                except Exception as e:
                    print(f"  ERR {slug}: {e}")
                    missing += 1
            else:
                print(f"  SKIP {slug}: unknown image source — {name}")
                missing += 1
        else:
            print(f"  MISS {slug}: no image at row {row} — {name}")
            missing += 1

    print(f"\nDone: {extracted} extracted, {missing} missing, total {len(agents)}")
    wb.close()


if __name__ == "__main__":
    main()
